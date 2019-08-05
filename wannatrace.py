import openpyxl
from openpyxl import load_workbook as lw
import crawling as cr
from datetime import datetime

def tracing_sellingpoint(filename, url_fn) :
    # filename, url_fn = '주식시장조사용_yes.xlsx', '주식시장조사용_yes.txt'

    # 엑셀 파일의 url 란에 추적을 희망하는 도서 url 을 넣는다.
    # 원하는 파일 명을 넣고
    print("open xlsx.... %s" % (filename))
    # filename = 'C:/googledrive/재테크일반시장조사_py/재테크일반 시장조사.xlsx'
    # filename = 'C:/Users/ehdtj/OneDrive/PycharmProjects/재테크일반 시장조사.xlsx'

    if 'yes' in filename :
        # url_list = open(url_fn).read().split('\n')
        book = lw(filename)
        sheet1 = book['Sheet1']

        print("getting selling point.....")
        selling_point = [str(datetime.today().date())]

        for idx, row in enumerate(sheet1.iter_rows(min_row = 2, values_only=True)) :
            url = row[0]
            tmp_par = cr.makepar(url) # html 추출
            tmp_sp = int(tmp_par.find('span', 'gd_sellNum').get_text(' ', strip=True).split(' ')[2]) #판매지수
            title = tmp_par.find('h2', 'gd_name').text #제목
            if row[1] is None : sheet1.cell(row = idx+2, column = 2, value=title) #제목 없으면 삽입
            print(tmp_sp, ' : \t ', title)
            selling_point.append(tmp_sp)

    if 'al' in filename or 'al' in url_fn :

        # url_list = open(url_fn).read().split('\n')
        book = lw(filename)
        sheet1 = book['Sheet1']

        print("getting selling point.....")
        selling_point = [str(datetime.today().date())]

        for idx, row in enumerate(sheet1.iter_rows(min_row = 2, values_only=True)) :
            url = row[0]
            tmp_par = cr.makepar(url)
            tmp_sp = int(tmp_par.find(style='display:inline-block;').strong.text)
            title = tmp_par.find('a', 'Ere_bo_title').text
            if row[1] is None: sheet1.cell(row=idx + 2, column=2, value=title)  # 제목 없으면 삽입
            print(tmp_sp,' : \t ', title)
            selling_point.append(tmp_sp)

    print("\ninserting value to xlsx.....")

    col = sheet1.max_column + 1
    for r in range(1, len(selling_point) + 1):
        sheet1.cell(row=r, column=col, value=selling_point[r - 1])

    print("saving to xlsx.....")
    book.save(filename)

    print("%s finish!!\n" % (filename))
    book.close()

    variation_check(filename)

    return 0
def tracing_sellingpoint_col(filename) :
    # filename, url_fn = '주식시장조사용_yes.xlsx', '주식시장조사용_yes.txt'

    # 엑셀 파일의 url 란에 추적을 희망하는 도서 url 을 넣는다.
    # 원하는 파일 명을 넣고
    print("open xlsx.... %s" % (filename))
    # filename = 'C:/googledrive/재테크일반시장조사_py/재테크일반 시장조사.xlsx'
    # filename = 'C:/Users/ehdtj/OneDrive/PycharmProjects/재테크일반 시장조사.xlsx'

    if 'yes' in filename :

        # url_list = open(url_fn).read().split('\n')
        print("getting selling point.....")
        selling_point = [str(datetime.today().date())]

        book2 = lw(filename)
        sheet2 = book2['Sheet1']

        for idx, col in enumerate(sheet2.iter_cols(min_col = 2, values_only=True)) :
            url = col[0]
            tmp_par = cr.makepar(url) # html 추출
            tmp_sp = int(tmp_par.find('span', 'gd_sellNum').get_text(' ', strip=True).split(' ')[2]) #판매지수
            title = tmp_par.find('h2', 'gd_name').text #제목
            if col[1] is None : sheet2.cell(column = idx+2, row = 2, value=title) #제목 없으면 삽입
            print(tmp_sp, ' : \t ', title)
            selling_point.append(tmp_sp)

        row = sheet2.max_row + 1
        for c in range(1, len(selling_point) + 1):
            sheet2.cell(row=row, column=c, value=selling_point[c - 1])

    print("saving to xlsx.....")
    book2.save(filename)

    print("%s finish!!\n" % (filename))
    book2.close()

    variation_check_col(filename)

    return 0
def variation_check(filename) :
    book = lw(filename)
    sheet1 = book['Sheet1']
    print("last 3days values\n")
    s_list = []
    for row in sheet1.iter_cols(min_col=sheet1.max_column - 3, max_col=sheet1.max_column, values_only=True):
        print(row)
        s_list.append(list(row))

    # s_list2 = []
    # for list in s_list :
    #     tmp_list = []
    #     for k in range(-4,0) :
    #         tmp_list.append(list[k]-list[k+1])
    #     s_list2.append(tmp_list)

    # s_list3 = ['variation_2nd_latest']
    # for i in range(1, len(s_list[-1])):
    #     s_list3.append(s_list[-2][i] - s_list[-3][i])
    # print(s_list3, '\n\n')
    s_list2 = ['variation_latest']
    for i in range(1, len(s_list[-1])):
        if s_list[-2][i] is None :
            s_list2.append('0')
        else : s_list2.append(s_list[-1][i] - s_list[-2][i])
        # print(s_list[-1][i] - s_list[-2][i])
    print(s_list2,'\n')

    book.close()
def variation_check_col(filename) :
    book = lw(filename)
    sheet1 = book['Sheet1']
    print("last 3days values\n")
    s_list = []
    for row in sheet1.iter_rows(min_row=sheet1.max_row - 3, max_row=sheet1.max_row, values_only=True):
        print(row)
        s_list.append(list(row))

    s_list2 = ['variation_latest']
    for i in range(1, len(s_list[-1])):
        if s_list[-2][i] is None:
            s_list2.append('0')
        else:
            s_list2.append(s_list[-1][i] - s_list[-2][i])
        # print(s_list[-1][i] - s_list[-2][i])
    print(s_list2, '\n')

    book.close()
