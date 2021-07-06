import openpyxl
from openpyxl import load_workbook as lw
import crawling as cr
from datetime import datetime

# 엑셀 파일의 url 란에 추적을 희망하는 도서 url 을 넣는다.
# 원하는 파일 명을 넣고
print("open xlsx....")
# filename = 'C:/googledrive/재테크일반시장조사_py/재테크일반 시장조사.xlsx'
# filename = 'C:/Users/ehdtj/OneDrive/PycharmProjects/재테크일반 시장조사.xlsx'
filename = '세금 재테크 상식사전_tracing.xlsx'
book=lw(filename)
sheet1=book['Sheet1']

print("making url list....")
url_list = []
for tmp in sheet1['a'] :
    url_list.append(tmp.value)
del url_list[0]
print("number of url list : %d !" % len(url_list))

today=datetime.today()

print("getting selling point.....")
selling_point = ['%d-%d-%d' %(today.year,today.month,today.day)]
for url in url_list :
    print(url)
    tmp_par = cr.makepar(url)
    tmp_sp = tmp_par.find('span', 'gd_sellNum').get_text(' ',strip=True).split(' ')[2]
    print(tmp_sp)
    selling_point.append(tmp_sp)

def save_to_xl(selling_point) :
    print("inserting value to xlsx.....")
    col = sheet1.max_column + 1
    for r in range(1,len(selling_point)+1) :
        sheet1.cell(row=r, column=col, value=selling_point[r-1])

    print("saving to xlsx.....")
    book.save(filename)
    return 0

save_to_xl(selling_point)

# data = []
# for row in sheet1.rows:
#     tmp = []
#     for t in range(0, len(row)):
#         tmp.append(row[t].value)
#     print(tmp)
#     data.append(tmp)

