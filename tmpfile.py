import openpyxl
from openpyxl import load_workbook as lw
import crawling as cr
from datetime import datetime

filename, url_fn = '주식시장조사용_yes.xlsx', '주식시장조사용_yes.txt'

# 엑셀 파일의 url 란에 추적을 희망하는 도서 url 을 넣는다.
# 원하는 파일 명을 넣고
print("open xlsx.... %s\n*******************" % (filename))
# filename = 'C:/googledrive/재테크일반시장조사_py/재테크일반 시장조사.xlsx'
# filename = 'C:/Users/ehdtj/OneDrive/PycharmProjects/재테크일반 시장조사.xlsx'

if 'yes' in filename or 'yes' in url_fn:

    url_list = open(url_fn).read().split('\n')
    book = lw(filename)
    sheet1 = book['Sheet1']

    # print("making url list....")
    # url_list = []
    # for tmp in sheet1['f'] :
    #     url_list.append(tmp.value)
    # del url_list[0]
    # print("number of url list : %d !" % len(url_list))

    print("getting selling point.....")
    selling_point = [str(datetime.today().date())]
    for url in url_list:
        # print(url)
        tmp_par = cr.makepar(url)
        tmp_sp = int(tmp_par.find('span', 'gd_sellNum').get_text(' ', strip=True).split(' ')[2])
        print(tmp_sp, ' : \t ', tmp_par.find('h2', 'gd_name').text)
        selling_point.append(tmp_sp)

if 'al' in filename or 'al' in url_fn:

    url_list = open(url_fn).read().split('\n')
    book = lw(filename)
    sheet1 = book['Sheet1']

    # print("making url list....")
    # url_list = []
    # for tmp in sheet1['f'] :
    #     url_list.append(tmp.value)
    # del url_list[0]
    # print("number of url list : %d !" % len(url_list))

    print("getting selling point.....")
    selling_point = [str(datetime.today().date())]
    for url in url_list:
        # print(url)
        tmp_par = cr.makepar(url)
        tmp_sp = int(tmp_par.find(style='display:inline-block;').strong.text)
        print(tmp_sp, ' : \t ', tmp_par.find('a', 'Ere_bo_title').text)
        selling_point.append(tmp_sp)

print("\ninserting value to xlsx.....")
col = sheet1.max_column + 1
for r in range(1, len(selling_point) + 1):
    sheet1.cell(row=r, column=col, value=selling_point[r - 1])

print("saving to xlsx.....\n*****************")
book.save(filename)

print("%s finish!!\n\n" % (filename))

book = lw(filename)
sheet1 = book['Sheet1']
print("last 3days values\n")
s_list = []
for row in sheet1.iter_cols(min_col=sheet1.max_column - 3, max_col=sheet1.max_column, values_only=True):
    print(row)
    s_list.append(list(row))

s_list2 = ['variation']
for i in range(1,len(s_list[-1])):
    s_list2.append(s_list[-1][i] - s_list[-2][i])
print(s_list2)